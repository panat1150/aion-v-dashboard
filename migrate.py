#!/usr/bin/env python3
"""
One-time migration: AION_V_Backend_with_route.xlsx → Supabase
Run once locally before deploying to Streamlit Cloud.

Usage:
  pip install supabase openpyxl
  export SUPABASE_URL="https://xxx.supabase.co"
  export SUPABASE_KEY="service_role_key_here"
  python migrate.py
"""
import os, sys, shutil, tempfile
from datetime import date, datetime
import openpyxl
from supabase import create_client

BAT = 75.2

BASE  = os.path.dirname(os.path.abspath(__file__))
FILES = ["AION_V_Backend_with_route.xlsx", "AION_V_Backend.xlsx"]


def get_path():
    for f in FILES:
        p = os.path.join(BASE, f)
        if os.path.exists(p):
            return p
    raise FileNotFoundError("XLSX not found")


def d2s(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v) if v else ""


def load_charges():
    src = get_path()
    tmp = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(src, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        try: os.remove(tmp)
        except: pass
    ws = wb["CHARGE LOG"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[1] or not row[3]:
            continue
        odo  = int(row[1])
        dist = float(row[2] or 0)
        kwh  = float(row[3])
        rate = float(row[4] or 0)
        cost = round(kwh * rate, 2)
        eff  = round(kwh / dist * 100, 2) if dist > 0 else 0
        cpkm = round(cost / dist, 3) if dist > 0 else 0
        ds   = d2s(row[0])
        rows.append({
            "date": ds, "odo": odo, "dist": dist, "kwh": kwh,
            "rate": rate, "cost": cost, "eff": eff, "cpkm": cpkm,
            "station": str(row[8] or ""), "type": str(row[9] or ""),
            "notes": str(row[10] or ""), "month": ds[:7],
        })
    return rows


def load_obd():
    src = get_path()
    tmp = tempfile.mktemp(suffix=".xlsx")
    shutil.copy2(src, tmp)
    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
    finally:
        try: os.remove(tmp)
        except: pass
    ws = wb["OBD LOG"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        mcv, ncv = float(row[6] or 0), float(row[7] or 0)
        td, tc   = float(row[13] or 0), float(row[12] or 0)
        rows.append({
            "date": d2s(row[0]), "session": str(row[1] or ""),
            "soc": float(row[2] or 0), "soh": float(row[3] or 0),
            "pack_v": float(row[4] or 0), "current": float(row[5] or 0),
            "max_cell_v": mcv, "min_cell_v": ncv,
            "spread": round((mcv - ncv) * 1000, 2) if mcv and ncv else 0,
            "max_temp": float(row[9] or 0), "min_temp": float(row[10] or 0),
            "tc": tc, "td": td,
            "cycles": round(td / BAT, 1) if td > 0 else 0,
            "rt_eff": round(td / tc, 4) if tc > 0 else 0,
            "odo": int(row[14] or 0), "notes": str(row[15] or ""),
        })
    return rows


def main():
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_KEY")
    if not url or not key:
        print("ERROR: set SUPABASE_URL and SUPABASE_KEY env vars")
        sys.exit(1)

    client = create_client(url, key)

    print("Reading XLSX...")
    charges = load_charges()
    obds    = load_obd()
    print(f"  Charge sessions: {len(charges)}")
    print(f"  OBD snapshots:   {len(obds)}")

    print("Uploading to Supabase...")
    client.table("data_files").upsert({"key": "charge_log", "content": charges}).execute()
    client.table("data_files").upsert({"key": "obd_log",    "content": obds}).execute()
    print("Done.")


if __name__ == "__main__":
    main()
